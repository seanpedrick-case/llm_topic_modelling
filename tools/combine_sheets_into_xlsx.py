import os
from datetime import date, datetime
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from tools.config import OUTPUT_FOLDER
from tools.helper_functions import (
    clean_column_name,
    convert_reference_table_to_pivot_table,
    get_basic_response_data,
    load_in_data_file,
)


def add_cover_sheet(
    wb: Workbook,
    intro_paragraphs: list[str],
    model_name: str,
    analysis_date: str,
    analysis_cost: str,
    number_of_responses: int,
    number_of_responses_with_text: int,
    number_of_responses_with_text_five_plus_words: int,
    llm_call_number: int,
    input_tokens: int,
    output_tokens: int,
    time_taken: float,
    file_name: str,
    column_name: str,
    number_of_responses_with_topic_assignment: int,
    custom_title: str = "Cover Sheet",
):
    ws = wb.create_sheet(title=custom_title, index=0)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Write title
    ws["A1"] = "Large Language Model Topic analysis"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")

    # Add intro paragraphs
    row = 3
    for paragraph in intro_paragraphs:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1, value=paragraph)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 60  # Adjust height as needed
        row += 2

    # Add metadata
    meta_start = row + 1
    metadata = {
        "Date Excel file created": date.today().strftime("%Y-%m-%d"),
        "File name": file_name,
        "Column name": column_name,
        "Model name": model_name,
        "Analysis date": analysis_date,
        # "Analysis cost": analysis_cost,
        "Number of responses": number_of_responses,
        "Number of responses with text": number_of_responses_with_text,
        "Number of responses with text five plus words": number_of_responses_with_text_five_plus_words,
        "Number of responses with at least one assigned topic": number_of_responses_with_topic_assignment,
        "Number of LLM calls": llm_call_number,
        "Total number of input tokens from LLM calls": input_tokens,
        "Total number of output tokens from LLM calls": output_tokens,
        "Total time taken for all LLM calls (seconds)": round(float(time_taken), 1),
    }

    for i, (label, value) in enumerate(metadata.items()):
        row_num = meta_start + i
        ws[f"A{row_num}"] = label
        ws[f"A{row_num}"].font = Font(bold=True)

        cell = ws[f"B{row_num}"]
        cell.value = value
        cell.alignment = Alignment(wrap_text=True)
        # Optional: Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 75

    # Ensure first row cells are wrapped on the cover sheet
    for col_idx in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.alignment = Alignment(wrap_text=True, vertical="center")


def csvs_to_excel(
    csv_files: list[str],
    output_filename: str,
    sheet_names: list[str] = None,
    column_widths: dict = None,  # Dict of {sheet_name: {col_letter: width}}
    wrap_text_columns: dict = None,  # Dict of {sheet_name: [col_letters]}
    intro_text: list[str] = None,
    model_name: str = "",
    analysis_date: str = "",
    analysis_cost: str = "",
    llm_call_number: int = 0,
    input_tokens: int = 0,
    output_tokens: int = 0,
    time_taken: float = 0,
    number_of_responses: int = 0,
    number_of_responses_with_text: int = 0,
    number_of_responses_with_text_five_plus_words: int = 0,
    column_name: str = "",
    number_of_responses_with_topic_assignment: int = 0,
    file_name: str = "",
    unique_reference_numbers: list = [],
):
    if intro_text is None:
        intro_text = list()

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for idx, csv_path in enumerate(csv_files):
        # Use provided sheet name or derive from file name
        sheet_name = (
            sheet_names[idx]
            if sheet_names and idx < len(sheet_names)
            else os.path.splitext(os.path.basename(csv_path))[0]
        )
        df = pd.read_csv(csv_path)

        if sheet_name == "Original data":
            try:
                # Create a copy to avoid modifying the original
                df_copy = df.copy()
                # Insert the Reference column at position 0 (first column)
                df_copy.insert(0, "Reference", unique_reference_numbers)
                df = df_copy
            except Exception as e:
                print("Could not add reference number to original data due to:", e)

        ws = wb.create_sheet(title=sheet_name)

        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=True), start=1
        ):
            ws.append(row)

            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=col_idx)

                # Bold header row
                if r_idx == 1:
                    cell.font = Font(bold=True)

                # Set vertical alignment to middle by default
                cell.alignment = Alignment(vertical="center")

            # Apply wrap text if needed
            if wrap_text_columns and sheet_name in wrap_text_columns:
                for col_letter in wrap_text_columns[sheet_name]:
                    cell = ws[f"{col_letter}{r_idx}"]
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Freeze top row for all data sheets
        ws.freeze_panes = "A2"

        # Ensure all header cells (first row) are wrapped
        for col_idx in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Set column widths
        if column_widths and sheet_name in column_widths:
            for col_letter, width in column_widths[sheet_name].items():
                ws.column_dimensions[col_letter].width = width

    add_cover_sheet(
        wb,
        intro_paragraphs=intro_text,
        model_name=model_name,
        analysis_date=analysis_date,
        analysis_cost=analysis_cost,
        number_of_responses=number_of_responses,
        number_of_responses_with_text=number_of_responses_with_text,
        number_of_responses_with_text_five_plus_words=number_of_responses_with_text_five_plus_words,
        llm_call_number=llm_call_number,
        input_tokens=input_tokens,
        output_tokens=output_tokens,
        time_taken=time_taken,
        file_name=file_name,
        column_name=column_name,
        number_of_responses_with_topic_assignment=number_of_responses_with_topic_assignment,
    )

    wb.save(output_filename)

    print(f"Output xlsx summary saved as '{output_filename}'")

    return output_filename


###
# Run the functions
###
def collect_output_csvs_and_create_excel_output(
    in_data_files: List,
    chosen_cols: list[str],
    reference_data_file_name_textbox: str,
    in_group_col: str,
    model_choice: str,
    master_reference_df_state: pd.DataFrame,
    master_unique_topics_df_state: pd.DataFrame,
    summarised_output_df: pd.DataFrame,
    missing_df_state: pd.DataFrame,
    excel_sheets: str = "",
    usage_logs_location: str = "",
    model_name_map: dict = dict(),
    output_folder: str = OUTPUT_FOLDER,
    structured_summaries: str = "No",
):
    """
    Collect together output CSVs from various output boxes and combine them into a single output Excel file.

    Args:
        in_data_files (List): A list of paths to the input data files.
        chosen_cols (list[str]): A list of column names selected for analysis.
        reference_data_file_name_textbox (str): The name of the reference data file.
        in_group_col (str): The column used for grouping the data.
        model_choice (str): The LLM model chosen for the analysis.
        master_reference_df_state (pd.DataFrame): The master DataFrame containing reference data.
        master_unique_topics_df_state (pd.DataFrame): The master DataFrame containing unique topics data.
        summarised_output_df (pd.DataFrame): DataFrame containing the summarised output.
        missing_df_state (pd.DataFrame): DataFrame containing information about missing data.
        excel_sheets (str): Information regarding Excel sheets, typically sheet names or structure.
        usage_logs_location (str, optional): Path to the usage logs CSV file. Defaults to "".
        model_name_map (dict, optional): A dictionary mapping model choices to their display names. Defaults to {}.
        output_folder (str, optional): The directory where the output Excel file will be saved. Defaults to OUTPUT_FOLDER.
        structured_summaries (str, optional): Indicates whether structured summaries are being produced ("Yes" or "No"). Defaults to "No".

    Returns:
        tuple: A tuple containing:
            - list: A list of paths to the generated Excel output files.
            - list: A duplicate of the list of paths to the generated Excel output files (for UI compatibility).
    """

    if structured_summaries == "Yes":
        structured_summaries = True
    else:
        structured_summaries = False

    if not chosen_cols:
        raise Exception("Could not find chosen column")

    today_date = datetime.today().strftime("%Y-%m-%d")
    original_data_file_path = os.path.abspath(in_data_files[0])

    csv_files = list()
    sheet_names = list()
    column_widths = dict()
    wrap_text_columns = dict()
    short_file_name = os.path.basename(reference_data_file_name_textbox)
    reference_pivot_table = pd.DataFrame()
    reference_table_csv_path = ""
    reference_pivot_table_csv_path = ""
    unique_topic_table_csv_path = ""
    missing_df_state_csv_path = ""
    overall_summary_csv_path = ""
    number_of_responses_with_topic_assignment = 0

    if in_group_col:
        group = in_group_col
    else:
        group = "All"

    overall_summary_csv_path = output_folder + "overall_summary_for_xlsx.csv"

    if structured_summaries is True and not master_unique_topics_df_state.empty:
        print("Producing overall summary based on structured summaries.")
        # Create structured summary from master_unique_topics_df_state
        structured_summary_data = list()

        # Group by 'Group' column
        for group_name, group_df in master_unique_topics_df_state.groupby("Group"):
            group_summary = f"## {group_name}\n\n"

            # Group by 'General topic' within each group
            for general_topic, topic_df in group_df.groupby("General topic"):
                group_summary += f"### {general_topic}\n\n"

                # Add subtopics under each general topic
                for _, row in topic_df.iterrows():
                    subtopic = row["Subtopic"]
                    summary = row["Summary"]
                    # sentiment = row.get('Sentiment', '')
                    # num_responses = row.get('Number of responses', '')

                    # Create subtopic entry
                    subtopic_entry = f"**{subtopic}**"
                    # if sentiment:
                    #     subtopic_entry += f" ({sentiment})"
                    # if num_responses:
                    #     subtopic_entry += f" - {num_responses} responses"
                    subtopic_entry += "\n\n"

                    if summary and pd.notna(summary):
                        subtopic_entry += f"{summary}\n\n"

                    group_summary += subtopic_entry

            # Add to structured summary data
            structured_summary_data.append(
                {"Group": group_name, "Summary": group_summary.strip()}
            )

        # Create DataFrame for structured summary
        structured_summary_df = pd.DataFrame(structured_summary_data)
        structured_summary_df.to_csv(overall_summary_csv_path, index=False)
    else:
        # Use original summarised_output_df
        structured_summary_df = summarised_output_df
        structured_summary_df.to_csv(overall_summary_csv_path, index=None)

    if not structured_summary_df.empty:
        csv_files.append(overall_summary_csv_path)
        sheet_names.append("Overall summary")
        column_widths["Overall summary"] = {"A": 20, "B": 100}
        wrap_text_columns["Overall summary"] = ["B"]

    if not master_reference_df_state.empty:
        # Simplify table to just responses column and the Response reference number
        file_data, file_name, num_batches = load_in_data_file(
            in_data_files, chosen_cols, 1, in_excel_sheets=excel_sheets
        )
        basic_response_data = get_basic_response_data(
            file_data, chosen_cols, verify_titles="No"
        )
        reference_pivot_table = convert_reference_table_to_pivot_table(
            master_reference_df_state, basic_response_data
        )

        unique_reference_numbers = basic_response_data["Reference"].tolist()

        try:
            master_reference_df_state.rename(
                columns={"Topic_number": "Topic number"}, inplace=True, errors="ignore"
            )
            master_reference_df_state.drop(
                columns=["1", "2", "3"], inplace=True, errors="ignore"
            )
        except Exception as e:
            print("Could not rename Topic_number due to", e)

        number_of_responses_with_topic_assignment = len(
            master_reference_df_state["Response References"].unique()
        )

        reference_table_csv_path = output_folder + "reference_df_for_xlsx.csv"
        master_reference_df_state.to_csv(reference_table_csv_path, index=None)

        reference_pivot_table_csv_path = (
            output_folder + "reference_pivot_df_for_xlsx.csv"
        )
        reference_pivot_table.to_csv(reference_pivot_table_csv_path, index=None)

        short_file_name = os.path.basename(file_name)

    if not master_unique_topics_df_state.empty:

        master_unique_topics_df_state.drop(
            columns=["1", "2", "3"], inplace=True, errors="ignore"
        )

        unique_topic_table_csv_path = (
            output_folder + "unique_topic_table_df_for_xlsx.csv"
        )
        master_unique_topics_df_state.to_csv(unique_topic_table_csv_path, index=None)

    if unique_topic_table_csv_path:
        csv_files.append(unique_topic_table_csv_path)
        sheet_names.append("Topic summary")
        column_widths["Topic summary"] = {"A": 25, "B": 25, "C": 15, "D": 15, "F": 100}
        wrap_text_columns["Topic summary"] = ["B", "F"]
    else:
        print("Relevant unique topic files not found, excluding from xlsx output.")

    if reference_table_csv_path:
        if structured_summaries:
            print(
                "Structured summaries are being produced, excluding response level data from xlsx output."
            )
        else:
            csv_files.append(reference_table_csv_path)
            sheet_names.append("Response level data")
            column_widths["Response level data"] = {"A": 15, "B": 30, "C": 40, "H": 100}
            wrap_text_columns["Response level data"] = ["C", "G"]
    else:
        print("Relevant reference files not found, excluding from xlsx output.")

    if reference_pivot_table_csv_path:
        if structured_summaries:
            print(
                "Structured summaries are being produced, excluding topic response pivot table from xlsx output."
            )
        else:
            csv_files.append(reference_pivot_table_csv_path)
            sheet_names.append("Topic response pivot table")

            if reference_pivot_table.empty:
                reference_pivot_table = pd.read_csv(reference_pivot_table_csv_path)

            # Base widths and wrap
            column_widths["Topic response pivot table"] = {"A": 25, "B": 100}
            wrap_text_columns["Topic response pivot table"] = ["B"]

            num_cols = len(reference_pivot_table.columns)
            col_letters = [get_column_letter(i) for i in range(3, num_cols + 1)]

            for col_letter in col_letters:
                column_widths["Topic response pivot table"][col_letter] = 25

            wrap_text_columns["Topic response pivot table"].extend(col_letters)
    else:
        print(
            "Relevant reference pivot table files not found, excluding from xlsx output."
        )

    if not missing_df_state.empty:
        missing_df_state_csv_path = output_folder + "missing_df_state_df_for_xlsx.csv"
        missing_df_state.to_csv(missing_df_state_csv_path, index=None)

    if missing_df_state_csv_path:
        if structured_summaries:
            print(
                "Structured summaries are being produced, excluding missing responses from xlsx output."
            )
        else:
            csv_files.append(missing_df_state_csv_path)
            sheet_names.append("Missing responses")
            column_widths["Missing responses"] = {"A": 25, "B": 30, "C": 50}
            wrap_text_columns["Missing responses"] = ["C"]
    else:
        print("Relevant missing responses files not found, excluding from xlsx output.")

    new_csv_files = csv_files.copy()

    # Original data file
    original_ext = os.path.splitext(original_data_file_path)[1].lower()
    if original_ext == ".csv":
        csv_files.append(original_data_file_path)
    else:
        # Read and convert to CSV
        if original_ext == ".xlsx":
            if excel_sheets:
                df = pd.read_excel(original_data_file_path, sheet_name=excel_sheets)
            else:
                df = pd.read_excel(original_data_file_path)
        elif original_ext == ".parquet":
            df = pd.read_parquet(original_data_file_path)
        else:
            raise Exception(f"Unsupported file type for original data: {original_ext}")

        # Save as CSV in output folder
        original_data_csv_path = os.path.join(
            output_folder,
            os.path.splitext(os.path.basename(original_data_file_path))[0]
            + "_for_xlsx.csv",
        )
        df.to_csv(original_data_csv_path, index=False)
        csv_files.append(original_data_csv_path)

    sheet_names.append("Original data")
    column_widths["Original data"] = {"A": 20, "B": 20, "C": 20}
    wrap_text_columns["Original data"] = ["C"]
    if isinstance(chosen_cols, list) and chosen_cols:
        chosen_cols = chosen_cols[0]
    else:
        chosen_cols = str(chosen_cols) if chosen_cols else ""

    # Intro page text
    intro_text = [
        "This workbook contains outputs from the large language model topic analysis of open text data. Each sheet corresponds to a different CSV report included in the analysis.",
        f"The file analysed was {short_file_name}, the column analysed was '{chosen_cols}' and the data was grouped by column '{group}'."
        " Please contact the LLM Topic Modelling app administrator if you need any explanation on how to use the results."
        "Large language models are not 100% accurate and may produce biased or harmful outputs. All outputs from this analysis **need to be checked by a human** to check for harmful outputs, false information, and bias.",
    ]

    # Get values for number of rows, number of responses, and number of responses longer than five words
    number_of_responses = basic_response_data.shape[0]
    # number_of_responses_with_text = basic_response_data["Response"].str.strip().notnull().sum()
    number_of_responses_with_text = (
        basic_response_data["Response"].str.strip().notnull()
        & (basic_response_data["Response"].str.split().str.len() >= 1)
    ).sum()
    number_of_responses_with_text_five_plus_words = (
        basic_response_data["Response"].str.strip().notnull()
        & (basic_response_data["Response"].str.split().str.len() >= 5)
    ).sum()

    # Get number of LLM calls, input and output tokens
    if usage_logs_location:
        try:
            usage_logs = pd.read_csv(usage_logs_location)

            relevant_logs = usage_logs.loc[
                (
                    usage_logs["Reference data file name"]
                    == reference_data_file_name_textbox
                )
                & (
                    usage_logs[
                        "Large language model for topic extraction and summarisation"
                    ]
                    == model_choice
                )
                & (
                    usage_logs[
                        "Select the open text column of interest. In an Excel file, this shows columns across all sheets."
                    ]
                    == (chosen_cols[0] if isinstance(chosen_cols, list) and chosen_cols else chosen_cols)
                ),
                :,
            ]

            llm_call_number = sum(relevant_logs["Total LLM calls"].astype(int))
            input_tokens = sum(relevant_logs["Total input tokens"].astype(int))
            output_tokens = sum(relevant_logs["Total output tokens"].astype(int))
            time_taken = sum(
                relevant_logs["Estimated time taken (seconds)"].astype(float)
            )
        except Exception as e:
            print("Could not obtain usage logs due to:", e)
            usage_logs = pd.DataFrame()
            llm_call_number = 0
            input_tokens = 0
            output_tokens = 0
            time_taken = 0
    else:
        print("LLM call logs location not provided")
        usage_logs = pd.DataFrame()
        llm_call_number = 0
        input_tokens = 0
        output_tokens = 0
        time_taken = 0

    # Create short filename:
    model_choice_clean_short = clean_column_name(
        model_name_map[model_choice]["short_name"],
        max_length=20,
        front_characters=False,
    )
    # Extract first column name as string for cleaning and Excel output
    chosen_col_str = chosen_cols[0] if isinstance(chosen_cols, list) and chosen_cols else str(chosen_cols) if chosen_cols else ""
    in_column_cleaned = clean_column_name(chosen_col_str, max_length=20)
    file_name_cleaned = clean_column_name(
        file_name, max_length=20, front_characters=True
    )

    # Save outputs for each batch. If master file created, label file as master
    file_path_details = (
        f"{file_name_cleaned}_col_{in_column_cleaned}_{model_choice_clean_short}"
    )
    output_xlsx_filename = (
        output_folder
        + file_path_details
        + ("_structured_summaries" if structured_summaries else "_topic_analysis")
        + ".xlsx"
    )

    xlsx_output_filename = csvs_to_excel(
        csv_files=csv_files,
        output_filename=output_xlsx_filename,
        sheet_names=sheet_names,
        column_widths=column_widths,
        wrap_text_columns=wrap_text_columns,
        intro_text=intro_text,
        model_name=model_choice,
        analysis_date=today_date,
        analysis_cost="",
        llm_call_number=llm_call_number,
        input_tokens=input_tokens,
        output_tokens=output_tokens,
        time_taken=time_taken,
        number_of_responses=number_of_responses,
        number_of_responses_with_text=number_of_responses_with_text,
        number_of_responses_with_text_five_plus_words=number_of_responses_with_text_five_plus_words,
        column_name=chosen_col_str,
        number_of_responses_with_topic_assignment=number_of_responses_with_topic_assignment,
        file_name=short_file_name,
        unique_reference_numbers=unique_reference_numbers,
    )

    xlsx_output_filenames = [xlsx_output_filename]

    # Delete intermediate csv files
    for csv_file in new_csv_files:
        os.remove(csv_file)

    return xlsx_output_filenames, xlsx_output_filenames
